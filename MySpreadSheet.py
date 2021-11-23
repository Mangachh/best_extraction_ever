from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import Messages
import string_compares
import file_constants

# TODO: Cambiar clase y separar lo de guardar datos con la búsqueda de datos.
# La clase está quedando bonita, limpia y demás. Sé que hay varios puntos donde
# puede fallar el rendimiento, pero aún así está bastante bien, es python y no es un juego :D.
# Igualmente, mirar de separar un poco más y optimizar, pero antes:
# -Archivo con las cosas a buscar: esto tendría su gracia y lo haría bastante modular.


class MySpreadSheet:

    _POS_HC = 0
    _POS_BM = 1
    _POS_BIS = 2

    _POS_DEF_DICT = 0
    _POS_ABS_DICT = 1
    _POS_REL_DICT = 2

    def __init__(self, spread: Workbook):
        self.doc = spread
        self.absolute_search = [file_constants.CELL_SAMPLE_NAME]
        self.relative_search = ["Núm. Caixa $", "260/280 $", "260/230 $"]
        self.default_labels = ["BM Code", "Ext ID", "Ext value", "State"]
        return

    # todo: esto en otra clase -> Registrar Class
    def set_values_to_document(self, sheet_name: str, full_data: dict):
        '''Setea los valores en el documento'''
        print(Messages.MESS_WRITTING)

        # declaramos para poder pillar autocomplete, como guardamos los valores en el mismo sheet
        # que hemos abierto, pillamos las cosas de aquí
        sheet: Worksheet
        sheet_name = self.__get_real_sheet_name(sheet_name)
        sheet = self.doc[sheet_name]
        self.__populate_sheet(sheet, full_data)



    # todo: esto en otra clase -> Registrar Class
    def __print_header(self, sheet, row:int, column: int) -> None:
        index = 0
        tables = self.default_labels + self.absolute_search + self.relative_search
        for values in tables:
            sheet.cell(row, column + index).value = values
            index += 1


    def __populate_sheet(self, sheet: Workbook, full_data: dict) -> None:
        '''Rellena las celdas del sheet con los valores del diccionario.
           Tiene el control de errores para tratar con las idisiosincrasias del programa'''
        # pillamos el valor buscar
        (row, col, isValue) = self.__get_row_column_by_value("buscar", sheet)
        if isValue:
            self.__print_header(sheet, row, col + 1)

        # por cada celda, que es key en full_data
        for cell in full_data:
            full = ""
            index = 1
            print("Full data: ", full_data)

            # por cada celda-user, miramos las 3 listas que henos metidos
            for values_list in full_data[cell]:
                # y por cada lista escribimos datos (neeeeat)
                for def_value in values_list:
                    sheet.cell(cell.row, cell.column + index).value = def_value
                    index += 1
                    full += str(def_value) + ", "
                    print("Escribiendo Valor: ", str(def_value))

            Messages.print_value_inside_border(full)
        return
        # vale, esto imprime los valores por columnas
        for data in full_data[cell]:

            # sumamos 1 porque el cell que pillamos es el del valor por defecto
            sheet.cell(cell.row, cell.column + index + 1).value = data
            print("Data written: ", data)
            index += 1


    # TODO: esto en otra clase -> Registrar Clase
    def __observation_message(self, value_to_check) -> str:
        '''
        Método para imprimir el mensaje de observación. Depende de la etiqueta que tenga aquello que se le pasa
        imprimirá una etiqueta u otra. Lo que le solemos pasar es el ng,
        :param value_to_check: mira q
        :return:
        '''
        observation_mess =""
        if value_to_check == -1 or value_to_check is None:
            observation_mess = file_constants.NO_VOL_NAME
        elif value_to_check < 0:
            observation_mess = file_constants.POC_NAME_FULL
        elif value_to_check == 0:
            observation_mess = file_constants.SALIVA_EPI
        else:
            observation_mess = "Succes" # lo ponemos para debuggear
        return observation_mess

    def get_ids_form_sheet_name(self, sheet_name: str, cell_value: str) -> (list, bool):
        '''Pilla los valores que hemos dado desde una hoja a la que hemos dado el nombre.
           como nombre usaremos por defecto file_constatns.SHEET_NAME, en vez de hacerlo
           como constante, lo usamos como variable en la función, por si hay que cambiarlo'''
        name = self.__get_real_sheet_name(sheet_name)
        try:
            sheet = self.doc[name]
        except Exception as e:
            return None, False

        # si no hay nombre, devolvemos lista vacía y un false
        if name == file_constants.NO_NAME:
            return None, False

        # y aquí rellenamos
        valid_cells = []
        (row, col, isValue) = self.__get_row_column_by_value(cell_value, sheet)
        row += 1
        print("Columna de búsqueda = " + cell_value)
        for sheet_row in sheet.iter_rows(min_row=row, min_col=col, max_col=col):
            if sheet_row is not None:
                valid_cells.append(sheet_row)

        print("Celdas con datos: ", valid_cells)
        return valid_cells, True

    def __get_real_sheet_name(self, sheet_name: str) -> str:
        '''devuelve el nombre verdadero de la hoja hace una comparación entre el nombre de la hoja en minúscula
           y el que usamos nosotros en minúscula para luego devolver el nombre verdadero a partir del que hemos dado
           ej: nombre_hoja = "HoLa"; nombre del argumento = "Hola"
           así, esto devuelve  "HoLa"'''
        real_name = file_constants.NO_NAME

        for name in self.doc.sheetnames:
            if string_compares.equal_strings_not_case(name, sheet_name):
                return name

        return real_name

    def __get_row_column_by_value(self, cell_value: str, sheet) -> (int, int, bool):
        '''devuelve las coordenadas de fila y la columna que tenga el valor que hemos pasado.
        OJU! si hay varias celdas con el mismo valor, devuelve la primera de ellas
        :param cell_value: el valor de la celda
        :param sheet: hoja a mirar
        :return: fila, columna y si tiene valor
        '''
        col = 0
        row = 0
        is_value = False

        # por cada tupla de celdas en una sheet (cell_tuple)
        for cell_tp in sheet.rows:

            # por cada celda singular dentro de la tupla (cell_singular)
            for cell_sn in cell_tp:
                name = str(cell_sn.value).lower()
                if name == cell_value:
                    # si los nombres coinciden, pasamos las coordenadas
                    return cell_sn.row, cell_sn.column, True

        # también devolvemos la columna
        return row, col, is_value

    def get_complete_list(self, sheet, ids: list) -> dict:
        '''
        Toma como parametro la lista completa de id's y creamos un diccionario con todos los valores donde
        la clave es la celda que nos interesa. Aquí es donde, por ejemplo, si queremos que sea más
        procedural habrá que tocar. Seguramente en otro tipo de método será mejor.
        TODO: la idea es tener un txt con los nombres que queremos añadir sin el E1, etc
        :param sheet:
        :param ids:
        :return:
        '''
        final_dict = {}

        # pillamos las columnas por defect
        def_columns = self._default_columns(sheet)

        # ahora el número de las columnas donde estan las extracciones
        extract_cols = self.__get_all_ng_columns(sheet, file_constants.CELL_NG_NAME)


        # ahora, por cada fila de la columna del esquimot, miramos si el valor es igual a cualquiera de la lista
        # sumamos uno al row porque doy por hecho que la primera fila es la del título.
        # lo hacemos al revés, por cada valor de la lista miramos SI está en el esquimot
        for user_id in ids:
            for cells in sheet.iter_rows(min_row=1, min_col=def_columns[self._POS_HC], max_col=def_columns[self._POS_HC]):
                if cells[0].value is None:
                    continue

                if cells[0].value == user_id[0].value:
                    # miramos si tiene "BM Bis Code" para saber si está repetido
                    bis_cell = sheet.cell(cells[0].row, def_columns[self._POS_BIS])
                    if bis_cell.value is not None:
                        # si esta en el dicc, comprobamos el valor de la extracción
                        # TODO meter un tipo de constantes para esto o algo así, estaría bien porque no me entero
                        if user_id[0] in final_dict and final_dict[user_id[0]][3] > ext_value:
                            continue

                    # declaramos celdas para bm, sample y si hay bisBM (así tenemos autocomplete)
                    bm_cell: Cell
                    sam_cell: Cell
                    bis_cell: Cell

                    # pillamos el bm y el tipo de extracción
                    bm_cell = sheet.cell(cells[0].row, def_columns[self._POS_BM])
                    # sam_cell = sheet.cell(cells[0].row, samCol)

                    # ahora hacemos una tupla, pillamos la mejor extracción y sacamos el identificadr ("e1, e2, etc")
                    (ext_name, ext_value) = self.__get_best_extraction(sheet, cells[0].row, extract_cols)

                    # como los datos del excel no están unificados, pillamos el valor de la extracción para escribir
                    # un mensaje
                    message = self.__observation_message(ext_value)
                    default_tuple = (bm_cell.value, ext_name, ext_value, message)



                    # aquí metemos los fijos que no dependen del identificador
                    absolute_values = []
                    relative_values = []

                    absolute_values = self._get_absolute_values(sheet, cells[0].row)

                    #ahora los que viene con identificador, primero lo sacamos
                    ext_identifier = string_compares.regular_substring('e[0-9]', ext_name.lower())

                    if ext_identifier != "":
                        print("Identifier:", ext_identifier)
                        relative_values = self._get_relative_values(sheet, cells[0].row, ext_identifier)
                    else:
                        print("Indentifier: None")

                    # sumamos tuples para impresion
                    final_tuple = tuple(list(default_tuple) + absolute_values + relative_values)
                    Messages.print_value_inside_border(str(user_id[0].value) + ', '.join(str(i) for i in final_tuple))

                    # ok, por cada user hay una con los valores que nos interesan, así será más fácil organizar
                    values_list = []
                    values_list.insert(self._POS_DEF_DICT, default_tuple)
                    values_list.insert(self._POS_ABS_DICT, absolute_values)
                    values_list.insert(self._POS_REL_DICT, relative_values)
                    final_dict[user_id[0]] = values_list

        return final_dict


    def _default_columns(self, sheet) -> list:
        '''
        Devuelve una lista con las columnas comunes de los datos. Es decir, pilla la columna donde estan los
        BM Code, los HC code y los HC bis. Es más fácil operar de esta manera y limpiamos un poco en las
        funciones más grandes.
        TODO: en algún momento separemos clases, así que, bueno...
        :return: Lista con las columnas de BM, HC y BIS HC
        '''
        colums = []
        # primero, pillamos la columna bm
        (row, bmCol, isTrue) = self.__get_row_column_by_value(file_constants.CELL_BM_NAME, sheet)

        # ahora pillamos la columna del esquimot
        (row, hcCol, isTrue) = self.__get_row_column_by_value(file_constants.CELL_HC_NAME, sheet)

        # ahora pillamos la columna del sample
        (row, bisCol, isTrue) = self.__get_row_column_by_value(file_constants.CELL_BIS_NAME, sheet)

        # ahora columna bis
        (row, bisCol, isTrue) = self.__get_row_column_by_value(file_constants.CELL_BIS_NAME, sheet)
        colums.insert(MySpreadSheet._POS_HC, hcCol)
        colums.insert(MySpreadSheet._POS_BM, bmCol)
        colums.insert(MySpreadSheet._POS_BIS, bisCol)

        return colums

    def _get_absolute_values(self, sheet, id_row) -> list:

        absolute_values = []
        for field in self.absolute_search:
            (row, col, isValue) = self.__get_row_column_by_value(field, sheet)
            print("Absolute column name: ", field)
            value = sheet.cell(id_row, col).value
            print("Value: ", value)
            absolute_values.append(value)

        return absolute_values

    def _get_relative_values(self, sheet, id_row, identifier: str) -> list:
        relative_values = []

        for field in self.relative_search:
            full_name = identifier.join(field.split("$"))
            print("Relative column name: ", full_name)
            (row, col, isValue) = self.__get_row_column_by_value(full_name.lower(), sheet)
            value = sheet.cell(id_row, col).value
            print("Value: ", value)
            relative_values.append(value)

        return relative_values

    def __get_best_extraction(self, sheet, row: int, columns: list) -> (str, int):
        '''
        Devuelve la mejor extracción de las posibles, identificador y valor(e1, 12)
        OJUUU: este método es propenso a bugs por todas las excepciones que hay en las filas ya que hay números, texto
        y la nada mezclados
        :param sheet: hoja donde se hacen los cambios
        :param row: fila donde estan las extracciones
        :param columns: lista con las columnas donde se encuentras las extracciones
        :return: etiqueta extracción + valor
        '''
        # iniciamos tupla, así es más facil
        (name, best_value) = (file_constants.NO_VOL_NAME, file_constants.NO_VOL)
        index = 0
        #por cada columna en la que pillemos algo (oju, esto lo podriamos hacer de otro modo pero lo dejamos así)
        for col in columns:
            index += 1
            tempCell = sheet.cell(row, col)
            print(f"Value {index}: {tempCell.value}")
            # si no hay valor seguimos.
            # realmente antes había un break y parábamos, pero prefiero continuar
            # para pillar algún error, en plan, extracción e2 no tiene valor pero e3 sí
            if tempCell.value is None:
                continue
            elif type(tempCell.value) is str:
                if best_value == file_constants.NO_VOL and string_compares.file_constants.SALIVA_EPI.lower() in tempCell.value.lower():
                    (name, best_value) = file_constants.CELL_NG_NAME + str(index), file_constants.EPI_VOL
                continue
            elif sheet.cell(row, col - 1).value is not None:
                # pillamos la columna anterior al "ng/ul E1" que debería ser "VOL. NO DISPONIBLE"
                best_value_name = sheet.cell(row, col - 1).value

                # si no es texto, da igual y seguimos, esta comprobación es importante porque, como he dicho arriba
                # a veces hay números u otras cosas porque no funca bien la columna
                if type(best_value_name) is str:
                    best_value_name = best_value_name.lower()
                    # si la cadena incluye "poc" comprobamos valores
                    # metemos como absoluto el valor
                    if file_constants.POC_NAME.lower() in best_value_name:
                        if tempCell.value > abs(best_value):
                            (name, best_value) = file_constants.CELL_NG_NAME + str(index), - tempCell.value
                        #end
                    #end
                #end

            elif tempCell.value > best_value:
                (name, best_value) = file_constants.CELL_NG_NAME + str(index), tempCell.value

        print(f"Best value: {best_value}")
        return name, best_value

    def __get_all_ng_columns(self, sheet, name) -> list:
        '''
        pilla los index de todas las columans que tienen el nombre seleccionado
        :param sheet: hoja donde pillamos los datos, cambiar
        :param name: nombre lo que nos interesa
        :return: lista con los índices
        '''
        index = file_constants.EXTRACTION_FIRST
        has_data = True
        columns = []

        while has_data and index < file_constants.MAX_VARS:
            cell_value = name + str(index)
            print("data: ", cell_value)
            (row, col, hasData) = self.__get_row_column_by_value(cell_value, sheet)
            index += 1
            if hasData:
                columns.append(col)

        return columns
